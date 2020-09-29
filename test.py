

from csv import reader
import pandas as pd
from openpyxl import load_workbook

wb2 = load_workbook('test.xlsx')
worksheet1 = wb2['Sheet1']
for row in worksheet1.iter_rows():
    first_name = row[0].value
    last_name = row[1].value
    email = row[2].value
