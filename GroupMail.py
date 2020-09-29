#!/usr/bin/env python

__author__ = "Alyssa Shen"
__email__ = "alyssa.xinyun@gmail.com"
__description__ = "Send email to a group of people by parsing the excel file"

import csv
import sys
import smtplib
from openpyxl import load_workbook
from email.mime.text import MIMEText
from optparse import OptionParser


DELIMITER = ','
QUOTECHAR = '"'
TITLE_FIX = ""

SMTP_SERVER = "smtp.gmail.com:587"
FROM_EMAIL = "alyssa.xinyun@gmail.com"
FROM_PASSWORD = "password"
TO_DOMAIN = "@gmail.com"
CC_EMAIL = "alyssa.xinyun@gmail.com"

SUBJECT = "Invitation from Alyssa"
CONTENT_PREFIX = """
Invitation:
"""
CONTENT_POSTFIX = """
Regards,
Alyssa

"""


class GroupMail(object):

    title = []
    credit = []
    report = {}

    option_dic = {
        'title': title.append,
        'credit': credit.append
    }

    def __init__(self, *args, **kwargs):
        self.send = kwargs.get('send', False)
        self.no_confirm = kwargs.get('no_confirm', False)
        if self.send:
            self.server = smtplib.SMTP(SMTP_SERVER)
            self.server.starttls()
            self.server.login(FROM_EMAIL, FROM_PASSWORD)

    def execute(self, filename):
        wb2 = load_workbook(filename)
        worksheet1 = wb2['Sheet1']
        for row in worksheet1.iter_rows():
            first_name = row[0].value
            last_name = row[1].value
            email = row[2].value
            self.render_line(row)

    def render_line(self, row):
        content = "Hi, " + row[0].value + ":\n"
        target = row[2].value
        print("email address:", target)
        content += "test"

        content += CONTENT_POSTFIX
        if self.send is True:
            self.send_mail(target=target, content=content)
        else:
            print(content)

    def send_mail(self, target, title=None, content=None):
        msg = MIMEText(content)
        msg['From'] = FROM_EMAIL
        msg['To'] = target
        msg["Cc"] = CC_EMAIL
        msg['Subject'] = SUBJECT
        if self.no_confirm is False:
            print(msg.as_string())
            var = input("Send mail? (yes/n): ")
            if var == 'yes':
                self.server.sendmail(FROM_EMAIL, [target] + [CC_EMAIL], msg.as_string())
            else:
                self.report[target] = False

    def __del__(self):
        self.server.quit()


def main():
    usage = "usage: groupmail.py file [options]"
    parser = OptionParser(usage=usage)
    parser.add_option("-s", "--send", action="store_true", default=False,
                      help="send email, if not set, just print e-mail content")
    parser.add_option("-n", "--no-confirm", action="store_true", default=False,
                      help="no ask for comfirmation of the content of each e-mail before send")
    (options, args) = parser.parse_args()
    print(options)
    if not args:
        parser.print_help()
        sys.exit(1)
    obj = GroupMail(send=options.send, no_confirm=options.no_confirm)
    obj.execute(args[0])


if __name__ == "__main__":
    main()
